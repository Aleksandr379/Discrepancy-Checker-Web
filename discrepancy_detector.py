from flask import Flask, request, render_template_string, abort, send_from_directory, url_for, current_app
import pandas as pd
import pdfplumber
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import io, os, uuid

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB limit

ALLOWED_EXTENSIONS = {'pdf', 'xlsx', 'xls'}
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

HTML = """<html><head><title>PO & Invoice Checker</title></head><body>
  <h1>Upload Purchase Order and Invoice Files</h1>
  {% if error %}
    <p style="color:red;"><strong>Error: {{ error }}</strong></p>
  {% endif %}
  <form action="/" method="post" enctype="multipart/form-data">
    Purchase Order (PDF or Excel): <input type="file" name="po_file" required><br><br>
    Invoice (PDF or Excel): <input type="file" name="invoice_file" required><br><br>
    <input type="submit" value="Upload and Check">
  </form>
  {% if excel_filename and pdf_filename %}
    <h2>Discrepancies Found:</h2>
    <a href="{{ url_for('download_excel', filename=excel_filename) }}">Download Excel Report</a><br>
    <a href="{{ url_for('download_pdf', filename=pdf_filename) }}">Download PDF Report</a><br>
  {% endif %}
</body></html>"""

COLUMN_MAP = {
    'po number': ['po number', 'po id', 'po detail', 'purchase order'],
    'invoice number': ['invoice number', 'inv no', 'invoice id'],
    'vendor': ['vendor', 'supplier', 'vendor name', 'supplier name'],
    'total amount': ['total amount', 'amount', 'total'],
    'quantity': ['quantity', 'qty'],
    'currency': ['currency', 'curr']
}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def normalize_columns(df):
    df_cols = [str(c).lower() for c in df.columns]
    new_cols = {}
    for std, variants in COLUMN_MAP.items():
        for v in variants:
            if v in df_cols:
                new_cols[v] = std
                break
    df.rename(columns=new_cols, inplace=True)
    return df

def read_excel(path):
    return pd.read_excel(path)

def read_pdf(path):
    text = ''
    with pdfplumber.open(path) as pdf:
        for p in pdf.pages:
            txt = p.extract_text()
            if txt:
                text += txt + '\n'
    lines = text.split('\n')
    data = [line.split() for line in lines if line.strip()]
    return pd.DataFrame(data)

def extract_data(path):
    lower = path.lower()
    if lower.endswith(('.xls', '.xlsx')):
        return read_excel(path)
    elif lower.endswith('.pdf'):
        return read_pdf(path)
    else:
        raise ValueError("Unsupported file type.")

def find_discrepancies(po_df, invoice_df):
    po_df = normalize_columns(po_df)
    invoice_df = normalize_columns(invoice_df)
    discrepancies = []
    for _, po_row in po_df.iterrows():
        po_number = po_row.get('po number')
        if not po_number:
            continue
        matched = invoice_df[invoice_df.get('po number') == po_number]
        if matched.empty:
            discrepancies.append({'PO Number': po_number, 'Issue': 'No matching invoice'})
        else:
            inv = matched.iloc[0]
            for col in ['total amount', 'vendor', 'quantity', 'currency']:
                if po_row.get(col) != inv.get(col):
                    discrepancies.append({
                        'PO Number': po_number,
                        'Field': col,
                        'PO Value': po_row.get(col),
                        'Invoice Value': inv.get(col)
                    })
    return pd.DataFrame(discrepancies) if discrepancies else pd.DataFrame([{'Message': 'No discrepancies found!'}])

def save_excel_with_watermark(df, path, watermark="Discrepancy Detector"):
    df.to_excel(path, index=False)
    wb = load_workbook(path)
    ws = wb.active
    ws['A1'] = watermark
    wb.save(path)

def add_pdf_watermark_from_df(df, path, watermark="Discrepancy Detector"):
    packet = io.BytesIO()
    c = canvas.Canvas(packet, pagesize=letter)
    textobj = c.beginText(40, 750)
    for row in df.values:
        textobj.textLine(" | ".join([str(v) for v in row]))
    c.drawText(textobj)
    c.saveState()
    c.setFont("Helvetica", 40)
    try:
        c.setFillAlpha(0.2)
    except Exception:
        pass
    c.translate(300, 400)
    c.rotate(30)
    c.drawCentredString(0, 0, watermark)
    c.restoreState()
    c.showPage()
    c.save()
    packet.seek(0)
    with open(path, 'wb') as f:
        f.write(packet.read())

@app.route("/", methods=["GET", "POST"])
def home():
    error = None
    excel_filename = pdf_filename = None
    try:
        if request.method == "POST":
            if 'po_file' not in request.files or 'invoice_file' not in request.files:
                abort(400, description="Missing file upload")
            po = request.files['po_file']
            inv = request.files['invoice_file']
            if po.filename == "" or inv.filename == "":
                abort(400, description="No file selected")
            if not allowed_file(po.filename) or not allowed_file(inv.filename):
                abort(400, description="Unsupported file type")

            po_fn = secure_filename(po.filename)
            inv_fn = secure_filename(inv.filename)
            po_fn = f"{uuid.uuid4().hex}_{po_fn}"
            inv_fn = f"{uuid.uuid4().hex}_{inv_fn}"

            po_path = os.path.join(app.config['UPLOAD_FOLDER'], po_fn)
            inv_path = os.path.join(app.config['UPLOAD_FOLDER'], inv_fn)
            po.save(po_path)
            inv.save(inv_path)

            po_df = extract_data(po_path)
            inv_df = extract_data(inv_path)
            discrepancies = find_discrepancies(po_df, inv_df)

            uid = uuid.uuid4().hex
            excel_filename = f"report_{uid}.xlsx"
            pdf_filename = f"report_{uid}.pdf"
            excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
            pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], pdf_filename)

            save_excel_with_watermark(discrepancies, excel_path)
            add_pdf_watermark_from_df(discrepancies, pdf_path)

    except Exception as e:
        app.logger.error("Server error", exc_info=True)
        error = str(e)

    return render_template_string(
        HTML,
        error=error,
        excel_filename=excel_filename,
        pdf_filename=pdf_filename
    )

@app.route("/download/excel/<path:filename>")
def download_excel(filename):
    # validate filename to prevent path traversal
    if '..' in filename or filename.startswith('/'):
        abort(403)
    directory = os.path.join(current_app.root_path, app.config['UPLOAD_FOLDER'])
    return send_from_directory(directory, filename, as_attachment=True)

@app.route("/download/pdf/<path:filename>")
def download_pdf(filename):
    if '..' in filename or filename.startswith('/'):
        abort(403)
    directory = os.path.join(current_app.root_path, app.config['UPLOAD_FOLDER'])
    return send_from_directory(directory, filename, as_attachment=True)

@app.errorhandler(413)
def handle_too_large(e):
    return "Uploaded file too large (limit exceeded).", 413

@app.errorhandler(Exception)
def handle_unexpected_error(e):
    app.logger.error("Server error", exc_info=True)
    return "An internal server error occurred. Please try again later.", 500

if __name__ == "__main__":
    app.run(debug=True)
